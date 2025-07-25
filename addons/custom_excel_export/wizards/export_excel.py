from odoo import models, fields
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

class ExportExcelWizard(models.TransientModel):
    _name = 'export.excel.wizard'
    _description = 'Export Excel Wizard'

    def export_to_excel(self):
        # Initialize workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Báo Giá"

        # Define styles
        font_title = Font(name='Times New Roman', size=20, bold=True)
        font_header = Font(name='Times New Roman', size=14, bold=True)
        font_normal = Font(name='Times New Roman', size=14)
        font_company = Font(name='Times New Roman', size=14, bold=True, italic=True)
        font_company_info = Font(name='Times New Roman', size=14, italic=True)
        font_bold_blue = Font(name='Times New Roman', size=16, bold=True, color='27B1FC')
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_summary = PatternFill(start_color='E6E2D3', end_color='E6E2D3', fill_type='solid')

        # Get sale order data
        sale_orders = self.env['sale.order'].browse(self._context.get('active_ids', []))
        current_row = 1

        for doc in sale_orders:
            # Company logo
            company = doc.company_id
            if company.logo:
                logo_data = base64.b64decode(company.logo)
                img = PILImage.open(BytesIO(logo_data))
                img = img.convert('RGB')
                img_width, img_height = img.size
                scale = 250 / img_width
                img = img.resize((250, int(img_height * scale)))
                img_io = BytesIO()
                img.save(img_io, format='PNG')
                img_io.seek(0)
                logo = Image(img_io)
                ws.add_image(logo, 'A1')
                ws.row_dimensions[1].height = int(img_height * scale * 0.75)
                ws.merge_cells(start_row=1, start_column=1, end_row=4, end_column=5)
                current_row = 5
            else:
                current_row = 1

            # Company information
            ws.cell(row=1, column=6).value = "CÔNG TY TNHH GIẢI PHÁP KỸ THUẬT Y TẾ MIỀN NAM"
            ws.cell(row=1, column=6).font = font_company
            ws.cell(row=1, column=6).alignment = align_right
            ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=10)

            ws.cell(row=2, column=6).value = "Phone: (+84) 932.760.599"
            ws.cell(row=2, column=6).font = font_company_info
            ws.cell(row=2, column=6).alignment = align_right
            ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=10)

            ws.cell(row=3, column=6).value = "sales@somed.vn"
            ws.cell(row=3, column=6).font = font_company_info
            ws.cell(row=3, column=6).alignment = align_right
            ws.merge_cells(start_row=3, start_column=6, end_row=3, end_column=10)

            ws.cell(row=4, column=6).value = "www.somed.vn"
            ws.cell(row=4, column=6).font = font_company_info
            ws.cell(row=4, column=6).alignment = align_right
            ws.merge_cells(start_row=4, start_column=6, end_row=4, end_column=10)

            # Bottom border for header
            for col in range(1, 11):
                ws.cell(row=4, column=col).border = Border(bottom=Side(style='thin'))

            if not company.logo:
                current_row = 5

            # Title: BẢNG BÁO GIÁ
            ws.cell(row=current_row, column=1).value = "BẢNG BÁO GIÁ"
            ws.cell(row=current_row, column=1).font = font_title
            ws.cell(row=current_row, column=1).alignment = align_center
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            current_row += 1

            # Quotation number and project name
            ws.cell(row=current_row, column=8).value = doc.name
            ws.cell(row=current_row, column=8).font = font_bold_blue
            ws.cell(row=current_row, column=8).alignment = align_right
            if doc.x_project_name:
                ws.cell(row=current_row, column=1).value = f"Dự án: {doc.x_project_name}"
                ws.cell(row=current_row, column=1).font = font_normal
                ws.cell(row=current_row, column=1).alignment = align_left
            current_row += 1

            # Customer and employee information
            emp = self.env['hr.employee'].search([('user_id', '=', doc.user_id.id)], limit=1)
            ws.append([""] * 10)
            current_row += 1

            ws.cell(row=current_row, column=1).value = "Khách hàng:"
            ws.cell(row=current_row, column=1).font = font_header
            ws.cell(row=current_row, column=2).value = doc.partner_id.name
            ws.cell(row=current_row, column=2).font = font_normal
            ws.cell(row=current_row, column=6).value = "Nhân viên kinh doanh:"
            ws.cell(row=current_row, column=6).font = font_header
            ws.cell(row=current_row, column=7).value = doc.user_id.name
            ws.cell(row=current_row, column=7).font = font_normal
            current_row += 1

            ws.cell(row=current_row, column=1).value = "Địa chỉ:"
            ws.cell(row=current_row, column=1).font = font_header
            address = (doc.partner_id.contact_address or '').replace(doc.partner_id.name or '', '').strip()
            ws.cell(row=current_row, column=2).value = address
            ws.cell(row=current_row, column=2).font = font_normal
            ws.cell(row=current_row, column=6).value = "Địa chỉ:"
            ws.cell(row=current_row, column=6).font = font_header
            emp_address = (emp.address_id.contact_address or '').replace(emp.address_id.complete_name or '', '').strip() if emp.address_id else ''
            ws.cell(row=current_row, column=7).value = emp_address or "213 Đường TL15, Khu Phố 3C, P.Thạnh Lộc, Q.12, TP HCM"
            ws.cell(row=current_row, column=7).font = font_normal
            current_row += 1

            contact_phone = doc.partner_contact_id.phone or doc.partner_id.phone or ''
            ws.cell(row=current_row, column=1).value = "Điện thoại:"
            ws.cell(row=current_row, column=1).font = font_header
            ws.cell(row=current_row, column=2).value = contact_phone
            ws.cell(row=current_row, column=2).font = font_normal
            ws.cell(row=current_row, column=6).value = "Điện thoại:"
            ws.cell(row=current_row, column=6).font = font_header
            ws.cell(row=current_row, column=7).value = emp.mobile_phone or ''
            ws.cell(row=current_row, column=7).font = font_normal
            current_row += 1

            contact_email = doc.partner_contact_id.email or doc.partner_id.email or ''
            ws.cell(row=current_row, column=1).value = "Email:"
            ws.cell(row=current_row, column=1).font = font_header
            ws.cell(row=current_row, column=2).value = contact_email
            ws.cell(row=current_row, column=2).font = font_normal
            ws.cell(row=current_row, column=6).value = "Email:"
            ws.cell(row=current_row, column=6).font = font_header
            ws.cell(row=current_row, column=7).value = doc.user_id.email or ''
            ws.cell(row=current_row, column=7).font = font_normal
            current_row += 1

            # Introduction
            ws.append([""] * 10)
            current_row += 1
            ws.cell(row=current_row, column=1).value = (
                "Lời đầu tiên Công ty TNHH giải pháp kỹ thuật Y tế Miền Nam xin gửi Quý khách hàng lời chúc sức khỏe và lời chào trân trọng nhất!"
            )
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1).font = font_normal
            ws.cell(row=current_row, column=1).alignment = align_left
            current_row += 1
            ws.cell(row=current_row, column=1).value = (
                "Công ty TNHH giải pháp kỹ thuật Y tế Miền Nam xin gửi Quý khách hàng bảng báo giá sản phẩm, vật tư theo yêu cầu từ Quý khách hàng, cụ thể như sau:"
            )
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1).font = font_normal
            ws.cell(row=current_row, column=1).alignment = align_left
            current_row += 1

            # Product table headers
            headers = ["STT", "Sản phẩm"]
            if doc.is_show_ma_sp:
                headers.append("Mã SP")
            headers.extend(["Thông số", "Xuất xứ", "Đơn vị", "Khối lượng"])
            if doc.is_show_chi_phi_nhan_cong:
                headers.append("Chi phí nhân công")
            headers.extend(["Đơn giá", "Thành tiền", "Ghi chú"])
            ws.append(headers)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col)
                cell.font = font_header
                cell.alignment = align_center
                cell.border = border
            current_row += 1

            # Product lines
            stt = 0
            sttLM = 0
            for line in doc.order_line:
                if line.display_type != 'line_section':
                    stt += 1
                    row_data = [stt, line.product_id.name]
                    if doc.is_show_ma_sp:
                        row_data.append(line.product_id.default_code or '')
                    row_data.extend([
                        line.x_thongso or '',
                        line.x_xuatxu or '',
                        line.product_uom.name,
                        line.product_uom_qty,
                    ])
                    if doc.is_show_chi_phi_nhan_cong:
                        row_data.append(line.x_chi_phi_nhan_cong or 0.0)
                    row_data.extend([
                        line.price_unit,
                        line.price_subtotal,
                        line.x_note or ''
                    ])
                    ws.append(row_data)
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = font_normal
                        cell.alignment = align_center if col != 2 else align_left
                        cell.border = border
                    current_row += 1
                else:
                    sttLM += 1
                    stt = 0
                    roman_num = self.env['sale.order'].int_to_roman(sttLM)
                    row_data = [roman_num, line.name] + [""] * (len(headers) - 2)
                    ws.append(row_data)
                    for col in range(1, len(headers) + 1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.font = font_header
                        cell.alignment = align_center if col == 1 else align_left
                        cell.border = border
                        cell.fill = fill_summary
                    current_row += 1

            # Tax and summary
            tax_details = {}
            for line in doc.order_line:
                for tax in line.tax_id:
                    base = tax_details.get(tax.name, {'base': 0.0, 'amount': 0.0})
                    tax_details[tax.name] = {
                        'base': base['base'] + line.price_subtotal,
                        'amount': base['amount'] + (line.price_subtotal * tax.amount / 100)
                    }

            colspan = 10 if doc.is_show_chi_phi_nhan_cong and doc.is_show_ma_sp else 9 if doc.is_show_chi_phi_nhan_cong or doc.is_show_ma_sp else 8
            ws.append(["Tổng chưa VAT"] + [""] * (colspan - 2) + [doc.amount_untaxed, ""] + [""] * (len(headers) - 2))
            for col in range(1, colspan + 1):
                ws.cell(row=current_row, column=col).font = font_header
                ws.cell(row=current_row, column=col).alignment = align_left
                ws.cell(row=current_row, column=col).fill = fill_summary
                ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=colspan).value = doc.amount_untaxed
            ws.cell(row=current_row, column=colspan).font = font_header
            ws.cell(row=current_row, column=colspan).alignment = align_center
            ws.cell(row=current_row, column=colspan).border = border
            current_row += 1

            for tax_name, details in tax_details.items():
                ws.append([f"VAT {tax_name}"] + [""] * (colspan - 2) + [details['amount'], ""] + [""] * (len(headers) - 2))
                for col in range(1, colspan + 1):
                    ws.cell(row=current_row, column=col).font = font_header
                    ws.cell(row=current_row, column=col).alignment = align_left
                    ws.cell(row=current_row, column=col).fill = fill_summary
                    ws.cell(row=current_row, column=col).border = border
                ws.cell(row=current_row, column=colspan).value = details['amount']
                ws.cell(row=current_row, column=colspan).font = font_header
                ws.cell(row=current_row, column=colspan).alignment = align_center
                ws.cell(row=current_row, column=colspan).border = border
                current_row += 1

            ws.append(["Tổng"] + [""] * (colspan - 2) + [doc.amount_total, ""] + [""] * (len(headers) - 2))
            for col in range(1, colspan + 1):
                ws.cell(row=current_row, column=col).font = font_header
                ws.cell(row=current_row, column=col).alignment = align_left
                ws.cell(row=current_row, column=col).fill = fill_summary
                ws.cell(row=current_row, column=col).border = border
            ws.cell(row=current_row, column=colspan).value = doc.amount_total
            ws.cell(row=current_row, column=colspan).font = font_header
            ws.cell(row=current_row, column=colspan).alignment = align_center
            ws.cell(row=current_row, column=colspan).border = border
            current_row += 1

            # Amount in words
            ws.append(["Số tiền bằng chữ: " + (doc.number_to_text(doc.amount_total) or "")])
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1).font = font_header
            ws.cell(row=current_row, column=1).alignment = align_center
            current_row += 1

            # Commercial conditions
            ws.append([""] * 10)
            current_row += 1
            ws.cell(row=current_row, column=1).value = "Điều kiện thương mại:"
            ws.cell(row=current_row, column=1).font = font_header
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            current_row += 1

            line_num = 1
            if doc.is_including_testing:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Báo giá đã bao gồm chi phí kiểm định"
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
                ws.cell(row=current_row, column=1).font = font_normal
                ws.cell(row=current_row, column=1).alignment = align_left
                current_row += 1
                line_num += 1

            ws.cell(row=current_row, column=1).value = f"{line_num}. Báo giá đã bao gồm Thuế VAT"
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1).font = font_normal
            ws.cell(row=current_row, column=1).alignment = align_left
            current_row += 1
            line_num += 1

            if doc.is_including_installation and doc.is_including_transport:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Báo giá đã bao gồm lắp đặt và vận chuyển"
            elif doc.is_including_installation:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Báo giá đã bao gồm lắp đặt"
            elif doc.is_including_transport:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Báo giá đã bao gồm vận chuyển"
            else:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Báo giá chưa bao gồm lắp đặt và vận chuyển"
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1).font = font_normal
            ws.cell(row=current_row, column=1).alignment = align_left
            current_row += 1
            line_num += 1

            if doc.x_estimated_delivery_time_id:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Thời gian giao hàng: {doc.x_estimated_delivery_time_id.name}"
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
                ws.cell(row=current_row, column=1).font = font_normal
                ws.cell(row=current_row, column=1).alignment = align_left
                current_row += 1
                line_num += 1

            if doc.x_warranty_duration_id:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Thời gian bảo hành: {doc.x_warranty_duration_id.name}"
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
                ws.cell(row=current_row, column=1).font = font_normal
                ws.cell(row=current_row, column=1).alignment = align_left
                current_row += 1
                line_num += 1

            ws.cell(row=current_row, column=1).value = f"{line_num}. Điều khoản thanh toán:"
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1).font = font_normal
            ws.cell(row=current_row, column=1).alignment = align_left
            current_row += 1
            ws.cell(row=current_row, column=1).value = doc.x_custom_payment_terms or ''
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=1).font = font_normal
            ws.cell(row=current_row, column=1).alignment = align_left
            current_row += 1
            line_num += 1

            if doc.x_payment_method_id:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Phương thức thanh toán: {doc.x_payment_method_id.name}"
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
                ws.cell(row=current_row, column=1).font = font_normal
                ws.cell(row=current_row, column=1).alignment = align_left
                current_row += 1
                line_num += 1

            if doc.x_delivery_location:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Địa điểm giao hàng: {doc.x_delivery_location}"
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
                ws.cell(row=current_row, column=1).font = font_normal
                ws.cell(row=current_row, column=1).alignment = align_left
                current_row += 1
                line_num += 1

            if doc.validity_date and doc.date_order:
                ws.cell(row=current_row, column=1).value = f"{line_num}. Bảng báo giá có hiệu lực trong vòng {doc.x_quote_valid_until} ngày, sau đó có thể được thay đổi mà không thông báo trước."
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
                ws.cell(row=current_row, column=1).font = font_normal
                ws.cell(row=current_row, column=1).alignment = align_left
                current_row += 1
                line_num += 1

            # Signatures
            ws.append([""] * 10)
            current_row += 1
            ws.cell(row=current_row, column=1).value = "Xác nhận của khách hàng\n(Ký tên, đóng dấu)"
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
            ws.cell(row=current_row, column=1).font = Font(name='Times New Roman', size=14, italic=True)
            ws.cell(row=current_row, column=1).alignment = align_center
            ws.cell(row=current_row, column=6).value = (
                f"{doc.formatted_date or ''}\n"
                "CÔNG TY TNHH GIẢI PHÁP KỸ THUẬT Y TẾ MIỀN NAM\n"
                "PHÒNG KINH DOANH\n\n"
                f"{doc.user_id.name or ''}"
            )
            ws.merge_cells(start_row=current_row, start_column=6, end_row=current_row, end_column=10)
            ws.cell(row=current_row, column=6).font = Font(name='Times New Roman', size=14, bold=True, italic=True)
            ws.cell(row=current_row, column=6).alignment = align_center
            current_row += 1

            # Adjust column widths
            column_widths = [10, 20, 15, 30, 15, 15, 15, 15, 15, 20]
            if doc.is_show_ma_sp:
                column_widths[2] = 15
            if doc.is_show_chi_phi_nhan_cong:
                column_widths.insert(7, 15)
            for i, width in enumerate(column_widths[:len(headers)], 1):
                ws.column_dimensions[get_column_letter(i)].width = width

        # Save Excel file
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create attachment
        file_data = base64.b64encode(output.read())
        attachment = self.env['ir.attachment'].create({
            'name': f'sale_order_{doc.name}.xlsx',
            'type': 'binary',
            'datas': file_data,
            'res_model': 'sale.order',
            'res_id': doc.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }